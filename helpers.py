import csv
import math
import re
import pandas as pd
import json
import openpyxl
import numpy as np

keys_to_values = {
    'k.f1': '3-3-3-9',
    'k.f2': '3-4-5-6',
    'k.f3': '4-4-4-6',
    'k.f4': '4-5-3-6',
    'k.f5': '5-3-4-6',
    'k.f6': '1-1-1-15',
    'k.f7': '4-4-3-7',
    'k.f8': '3-4-4-7',
    'k.f9': '4-3-4-7',
    'k.f10': '3-5-4-6',
    'k.f11': '4-3-5-6',
    'k.f12': '5-4-3-6',
    'k.f99': 'natar'
}


def read_config(filename):
    base_url, username, password, proxy_url, x_, y_ = None, None, None, None, None, None
    with open(filename, 'r') as file:
        for line in file:
            key, value = line.strip().split(';', 1)
            if key.upper() == "BASE_URL":
                base_url = value
            elif key.upper() == "USERNAME":
                username = value
            elif key.upper() == "PASSWORD":
                password = value
            elif key.upper() == "PROXY_URL":
                proxy_url = value
            elif key == "x":
                x_ = int(value)
            elif key == "y":
                y_ = int(value)
    return base_url, username, password, proxy_url, x_, y_


def extract_animals_counts(tile_text, animal_types):
    animals_counts = {animal_id: 0 for animal_id in animal_types}
    for animal_id in animal_types:
        match = re.search(fr'<i class="unit {animal_id}"></i><span class="value ">(\d+)</span>', tile_text)
        if match:
            animals_counts[animal_id] = int(match.group(1))
    return animals_counts


def oasis_type_identifier(tile_text):
    """
    Identifies the oasis type based on the resource bonus patterns in the tile's text.
    Returns the identifier (e.g., 'id1', 'id2', ...) corresponding to the resource bonus.
    """
    oasis_types = {
        '50 Wood': r"{a:r1} {a\.r1} 50%",
        '50 Clay': r"{a:r2} {a\.r2} 50%",
        '50 Iron': r"{a:r3} {a\.r3} 50%",
        '50 Crop': r"{a:r4} {a\.r4} 50%",
        'Wood Crop': r"{a:r1} {a\.r1} 25%<br />{a:r4} {a\.r4} 25%",
        'Clay Crop': r"{a:r2} {a\.r2} 25%<br />{a:r4} {a\.r4} 25%",
        'Iron Crop': r"{a:r3} {a\.r3} 25%<br />{a:r4} {a\.r4} 25%",
        'Wood': r"{a:r1} {a\.r1} 25%",
        'Clay': r"{a:r2} {a\.r2} 25%",
        'Iron': r"{a:r3} {a\.r3} 25%",
        'Crop': r"{a:r4} {a\.r4} 25%",
    }
    for id_, pattern in oasis_types.items():
        if re.search(pattern, tile_text):
            return id_
    return 'unknown'


def find_oasis(api_response_data, oasis_output_path, executed_coords_output_path):
    animal_types = {
        'u31': 'Rats',
        'u32': 'Spiders',
        'u33': 'Snakes',
        'u34': 'Bats',
        'u35': 'Wild Boars',
        'u36': 'Wolves',
        'u37': 'Bears',
        'u38': 'Crocodiles',
        'u39': 'Tigers',
        'u40': 'Elephants',
    }

    executed_coords = set()
    try:
        with open(executed_coords_output_path, 'r', encoding='utf-8') as file:
            executed_coords = {line.strip() for line in file.readlines()}
    except FileNotFoundError:
        pass  # Dosya bulunamazsa boş bir set ile devam edin

    new_executed_coords = []

    with open(oasis_output_path, 'a', encoding='utf-8') as oasis_file:
        for tile in api_response_data['tiles']:
            if '{k.fo}' in tile.get('title', '') and '{k.vt}' not in tile.get('title', ''):  # Vaha kontrolü
                x = tile['position']['x']
                y = tile['position']['y']
                coord_key = f'{x};{y}'

                if coord_key not in executed_coords:
                    executed_coords.add(coord_key)
                    new_executed_coords.append(coord_key)

                    oasis_idx = oasis_type_identifier(tile['text'])
                    animals_counts = extract_animals_counts(tile['text'], animal_types)

                    line_parts = [f'{x};{y};{oasis_idx}'] + [str(animals_counts[animal]) for animal in
                                                             sorted(animal_types.keys())]
                    line = ';'.join(line_parts)
                    oasis_file.write(f'{line}\n')
    with open(executed_coords_output_path, 'a', encoding='utf-8') as file:
        for coord in new_executed_coords:
            file.write(f'{coord}\n')


def write_response(responses_path, response_data):
    response_str = json.dumps(response_data, ensure_ascii=False, indent=4)
    with open(responses_path, 'a', encoding='utf-8') as file:
        file.write(response_str + "\n")


def remove_duplicates_from_executed_coordinates(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        unique_coords = set(file.readlines())

    with open(file_path, 'w', encoding='utf-8') as file:
        for coord in sorted(unique_coords):
            file.write(coord)


def calc_distance(x1, y1, x2, y2, map_size=200):
    dx = min(abs(x2 - x1), (2 * map_size + 1) - abs(x2 - x1))
    dy = min(abs(y2 - y1), (2 * map_size + 1) - abs(y2 - y1))
    return math.sqrt(dx ** 2 + dy ** 2)


def sort_and_rewrite_oases(oasis_path, myX, myY, sort_algo=1):
    oasis_list = []
    # algo1 means sorting with oases score algo0 means sort with a distance
    if sort_algo == 2:  # sort with from minimum distance
        with open(oasis_path, 'r') as file:
            for line in file:
                parts = line.strip().split(';')
                x, y = int(parts[0]), int(parts[1])
                distance = calc_distance(myX, myY, x, y)
                oasis_list.append((distance,) + tuple(parts))
            oasis_list.sort(key=lambda x_: x_[0])
        with open(oasis_path, 'w') as file:
            for item in oasis_list:
                file.write(';'.join(item[1:]) + '\n')
    if sort_algo == 1:
        with open(oasis_path, 'r') as file:
            for line in file:
                parts = line.strip().split(';')
                x, y = int(parts[0]), int(parts[1])
                score = calculate_scores(parts[3:], myX, myY, x, y)
                oasis_list.append((score,) + tuple(parts))
            oasis_list.sort(key=lambda x_: x_[0], reverse=True)
        with open(oasis_path, 'w') as file:
            for item in oasis_list:
                file.write(';'.join(item[1:]) + '\n')


def oasis_to_excel(oasis_path, base_url, current_x=0, current_y=0):
    column_names = ["X", "Y", "VAHA TİPİ", "Sıçan", "Örümcek", "Yılan", "Yarasa", "Domuz",
                    "Kurt", "Ayı", "Timsah", "Kaplan", "Fil", "distance", "oasis_score", "attack_link"]
    df = pd.read_csv(oasis_path, sep=";", header=None, names=column_names[:-3])
    df["distance"] = df.apply(lambda row: round(calc_distance(row["X"], row["Y"], current_x, current_y), 2), axis=1)
    df["oasis_score"] = df.apply(lambda row: calculate_scores(
        [row["Sıçan"], row["Örümcek"], row["Yılan"], row["Yarasa"], row["Domuz"], row["Kurt"],
         row["Ayı"], row["Timsah"], row["Kaplan"], row["Fil"], ],
        current_x, current_y, row["X"], row["Y"], ), axis=1)
    df["attack_link"] = ""

    excel_path = oasis_path.replace('.txt', '.xlsx')
    df.to_excel(excel_path, index=False)
    print(f"{excel_path} olarak kaydedildi.")
    create_oasis_attack_links(excel_path, base_url)


def calculate_scores(animals_count, myX, myY, dX, dY):
    animals_rewards = [160, 160, 160, 160, 320, 320, 480, 480, 480, 800]
    animal_powers = [20, 40, 60, 50, 33, 70, 200, 240, 250, 520]
    # if you want to just see where is the maximum rats and boars use it this ones
    animals_rewards2 = [1600, 1600, 1600, 160, 3200, 100, 100, 100, 100, 100]
    animal_powers2 = [20, 30, 160, 500, 33, 700, 2000, 2400, 2500, 5200]
    distance = calc_distance(myX, myY, dX, dY)
    # Ensure animals_count is an array of integers
    animals_count = np.array(animals_count, dtype=np.int64)
    total_power = np.sum(animals_count * np.array(animal_powers))
    total_reward = np.sum(animals_count * np.array(animals_rewards))
    numerator = total_reward
    denominator = total_power * distance
    if denominator == 0:
        return 0
    return round(numerator / denominator, 3)


def create_oasis_attack_links(oasis_excel_path, BASE_URL, attack_type=3, soldier_unit_no=4, soldier_number=100):
    wb = openpyxl.load_workbook(oasis_excel_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_col=16, max_row=ws.max_row):
        x = row[0].value
        y = row[1].value
        # link = f"{BASE_URL}/build.php?id=39&tt=2&x={x}&y={y}&troop[t{soldier_unit_no}]={soldier_number}" \
        # f"&c=3&gid=16&eventType={attack_type}"
        link = f"{BASE_URL}/karte.php?x={x}&y={y}"
        link_cell = row[15]
        link_cell.hyperlink = link
        link_cell.style = 'Hyperlink'  # Hücreyi hyperlink stilinde formatla
    wb.save(oasis_excel_path)
    print(f"Saldırı linkleri {oasis_excel_path} dosyasına eklendi.")


def sort_and_rewrite_valleys(valley_path, myX, myY):
    valley_list = []
    with open(valley_path, 'r') as file:
        for line in file:
            parts = line.strip().split(';')
            x, y = int(parts[0]), int(parts[1])
            distance = calc_distance(myX, myY, x, y)
            valley_list.append((distance,) + tuple(parts))
        valley_list.sort(key=lambda x_: x_[0])


def process_tiles(input_file_path, output_file_path):
    unique_coordinates = set()

    # `k.f99` kullanımı çıkarıldı. Sadece `{k.vt}` ve `{k.fX}` birlikte olanları eşleştir.
    pattern = re.compile(
        r'"position":\s*{\s*"x":\s*(-?\d+),\s*"y":\s*(-?\d+)\s*},.*?"title":\s*"\{k\.vt\} \{(k\.f\d+)}"',
        re.DOTALL)

    with open(input_file_path, 'r') as file:
        data_content = file.read()

    matches = pattern.findall(data_content)

    with open(output_file_path, 'w') as output_file:
        for x, y, kf_code in matches:
            coordinate_key = f"{x};{y}"
            if coordinate_key not in unique_coordinates:
                unique_coordinates.add(coordinate_key)
                # `keys_to_values` global değişkenini kullanarak eşleşme sağla.
                mapped_value = keys_to_values.get(kf_code, 'unknown')
                output_line = f"{x};{y};{mapped_value}\n"
                if x == "-178" and y == "53":
                    print(f"output_line {x}, {y}, {kf_code}")
                output_file.write(output_line)


# Fonksiyonun kullanımı burada gösterilmemiştir, kendi dosya yollarınızla çağırmanız gerekmektedir.


def find_crops(valley_file_path, oasis_file_path, crops_file_path):
    valley_interest = ['1-1-1-15', '3-3-3-9']
    valley_coords = []

    # valley.txt'den ilgilenilen valley koordinatlarını oku
    with open(valley_file_path, 'r') as valley_file:
        for line in valley_file:
            parts = line.strip().split(';')
            if parts[2] in valley_interest:
                valley_coords.append((int(parts[0]), int(parts[1]), parts[2]))

    crops = []

    # calculate %crops by reading oasis.txt
    with open(oasis_file_path, 'r') as oasis_file:
        oasis_lines = oasis_file.readlines()

    for x, y, valley_type in valley_coords:
        grain_totals = []
        for line in oasis_lines:
            parts = line.strip().split(';')
            ox, oy = int(parts[0]), int(parts[1])
            if x - 3 <= ox <= x + 3 and y - 3 <= oy <= y + 3:  # 3x3 çevre kontrolü
                grain_amount = 50 if '50 Crop' in parts[2] else 25 if 'Crop' in parts[2] else 0
                grain_totals.append(grain_amount)
        grain_totals.sort(reverse=True)
        total_grain = sum(grain_totals[:3])  # En yüksek 3 tahıl miktarını topla
        crops.append(f"{x};{y};{valley_type};{total_grain}")

    with open(crops_file_path, 'w') as crops_file:
        for crop in crops:
            crops_file.write(crop + '\n')


def crops_to_excel(crop_path, myX, myY, min_):
    crops_data = pd.read_csv(crop_path, header=None, names=['X', 'Y', 'Village Type', '%Crops'], sep=';')
    crops_data['Distance'] = crops_data.apply(lambda row: round(calc_distance(myX, myY, row['X'], row['Y']), 3), axis=1)
    filtered_crops = crops_data[
        (crops_data['%Crops'] >= min_) & (crops_data['Village Type'].isin(['3-3-3-9', '1-1-1-15']))]
    sorted_crops = filtered_crops.sort_values(by=['Distance'])
    sorted_crops.to_excel("crops.xlsx", index=False, columns=['X', 'Y', 'Village Type', '%Crops', 'Distance'])


def best_places_finder(oasis_file_path, best_places_file_path):
    oasis_data = []
    with open(oasis_file_path, 'r') as file:
        for line in file:
            data = line.strip().split(';')
            x = int(data[0])
            y = int(data[1])
            oasis_type = data[2]
            score = calculate_oasis_score(oasis_type)
            oasis_data.append((x, y, score))

    best_places = []
    for x in range(-200, 201):
        for y in range(-200, 201):
            place_score = calculate_place_score(x, y, oasis_data)
            best_places.append((x, y, place_score))

    best_places.sort(key=lambda x: x[2], reverse=True)

    with open(best_places_file_path, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['x', 'y', 'place_score'])
        for x, y, place_score in best_places:
            writer.writerow([x, y, place_score])

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Best Places'
    worksheet.append(['x', 'y', 'place_score'])
    for x, y, place_score in best_places:
        worksheet.append([x, y, place_score])

    workbook.save(best_places_file_path.replace('.txt', '.xlsx'))


def calculate_oasis_score(oasis_type):
    if oasis_type == 'Iron':
        return 4
    elif oasis_type == 'Clay':
        return 3
    elif oasis_type == 'Iron Crop' or oasis_type == 'Clay Crop':
        return 2
    else:
        return 1


def calculate_place_score(x, y, oasis_data):
    score = 0
    for oasis_x, oasis_y, oasis_score in oasis_data:
        distance = calc_distance(x, y, oasis_x, oasis_y)
        if distance <= 10:
            score += oasis_score
    return score
